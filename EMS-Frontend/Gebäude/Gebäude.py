# -*- coding: utf-8 -*-
import os
import sys
import pandas as pd
import csv
import importlib
Import = importlib.import_module("EMS-Backend.Classes.Import")
import warnings
warnings.filterwarnings("error")

from PyQt5 import *
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets

class Ui_Gebäude(QMainWindow):

    def __init__(self):
        super().__init__()

        #self.setObjectName("self")
        self.resize(770, 261)
        self.setWindowTitle("Gebäude")
        
        self.label = QtWidgets.QLabel(self)
        self.label.setGeometry(QtCore.QRect(10, 20, 111, 21))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label.setText("Grunddaten")
        self.label_2 = QtWidgets.QLabel(self)
        self.label_2.setGeometry(QtCore.QRect(160, 20, 121, 21))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_2.setText("Bauteildaten")
        self.label_3 = QtWidgets.QLabel(self)
        self.label_3.setGeometry(QtCore.QRect(540, 20, 51, 21))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_3.setText("Profil")
        self.line = QtWidgets.QFrame(self)
        self.line.setGeometry(QtCore.QRect(130, 0, 20, 531))
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.line_2 = QtWidgets.QFrame(self)
        self.line_2.setGeometry(QtCore.QRect(520, 0, 20, 531))
        self.line_2.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")

        
        self.label_6 = QtWidgets.QLabel(self)
        self.label_6.setGeometry(QtCore.QRect(10, 50, 121, 16))
        self.label_6.setObjectName("label_6")
        self.label_6.setText("Brutogeschossfläche [m²]")
        self.doubleSpinBox_BGF = QtWidgets.QDoubleSpinBox(self)
        self.doubleSpinBox_BGF.setGeometry(QtCore.QRect(10, 70, 121, 22))
        self.doubleSpinBox_BGF.setObjectName("doubleSpinBox_BGF")
        self.doubleSpinBox_BGF.setRange(0,10**9)
        self.label_7 = QtWidgets.QLabel(self)
        self.label_7.setGeometry(QtCore.QRect(10, 100, 121, 16))
        self.label_7.setObjectName("label_7")
        self.label_7.setText("Grundstücksfläche [m²]")
        self.doubleSpinBox_GF = QtWidgets.QDoubleSpinBox(self)
        self.doubleSpinBox_GF.setGeometry(QtCore.QRect(10, 120, 121, 22))
        self.doubleSpinBox_GF.setObjectName("doubleSpinBox_GF")
        self.doubleSpinBox_GF.setRange(0,10**9)
        self.label_8 = QtWidgets.QLabel(self)
        self.label_8.setGeometry(QtCore.QRect(10, 200, 121, 31))
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setWordWrap(True)
        self.label_8.setObjectName("label_8")
        self.label_8.setText("Wärmekapazität           [Wh/m²/K]")
        self.doubleSpinBox_cp_Gebaeude = QtWidgets.QDoubleSpinBox(self)
        self.doubleSpinBox_cp_Gebaeude.setGeometry(QtCore.QRect(10, 232, 121, 22))
        self.doubleSpinBox_cp_Gebaeude.setObjectName("doubleSpinBox_cp_Gebaeude")
        self.doubleSpinBox_cp_Gebaeude.setRange(0,10**9)
        self.label_9 = QtWidgets.QLabel(self)
        self.label_9.setGeometry(QtCore.QRect(10, 150, 121, 16))
        self.label_9.setObjectName("label_9")
        self.label_9.setText("Geschosshöhe [m]")
        self.doubleSpinBox_Geschosshoehe = QtWidgets.QDoubleSpinBox(self)
        self.doubleSpinBox_Geschosshoehe.setGeometry(QtCore.QRect(10, 170, 121, 22))
        self.doubleSpinBox_Geschosshoehe.setObjectName("doubleSpinBox_Geschosshoehe")
        self.doubleSpinBox_Geschosshoehe.setRange(0,100)

        #Tabelle mit Bauteilen
        self.tableView_BauteilDaten = QtWidgets.QTableWidget(self)
        self.tableView_BauteilDaten.setGeometry(QtCore.QRect(160, 50, 345, 192))
        self.tableView_BauteilDaten.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectItems)
        self.tableView_BauteilDaten.setObjectName("tableView_BauteilDaten")
        self.tableView_BauteilDaten.setColumnCount(4)
        self.tableView_BauteilDaten.setHorizontalHeaderLabels(["Bauteil", "Fläche [m²]", "U-Wert [W/m²K]", "Temp-KorrFaktor"])

        li_Bauteile = ["Wand","Boden","Dach","Fenster"]
        for i,Bauteil in enumerate(li_Bauteile):
            rowPosition = self.tableView_BauteilDaten.rowCount()
            self.tableView_BauteilDaten.insertRow(rowPosition)
            #Profilname und die beiden Verbrauchsdaten werden auf Read-Only gestellt
            flags = QtCore.Qt.ItemFlags()
            flags != QtCore.Qt.ItemIsEditable

            item = QtWidgets.QTableWidgetItem(Bauteil)
            item.setFlags(flags)
            self.tableView_BauteilDaten.setItem(rowPosition , 0, item)
            for i in range(1,4,1):
                self.tableView_BauteilDaten.setItem(rowPosition , i, QtWidgets.QTableWidgetItem(""))
            self.tableView_BauteilDaten.resizeColumnsToContents()

        #Profil Stuff (Laden,Speichern,löschen etc.)
        self.lineEdit_Profil = QtWidgets.QLineEdit(self)
        self.lineEdit_Profil.setGeometry(QtCore.QRect(540, 70, 121, 20))
        self.lineEdit_Profil.setObjectName("lineEdit_Profil")
        self.comboBox_SelectProfile = QtWidgets.QComboBox(self)
        self.comboBox_SelectProfile.setGeometry(QtCore.QRect(540, 120, 121, 22))
        self.comboBox_SelectProfile.setObjectName("comboBox_SelectProfile")
        self.label_4 = QtWidgets.QLabel(self)
        self.label_4.setGeometry(QtCore.QRect(540, 50, 101, 16))
        self.label_4.setObjectName("label_4")
        self.label_4.setText("Eingabe Profilname")
        self.label_5 = QtWidgets.QLabel(self)
        self.label_5.setGeometry(QtCore.QRect(540, 100, 81, 16))
        self.label_5.setObjectName("label_5")
        self.label_5.setText("Auswahl Profil")

        self.pushButton_SaveProfile = QtWidgets.QPushButton(self)
        self.pushButton_SaveProfile.setGeometry(QtCore.QRect(670, 70, 75, 23))
        self.pushButton_SaveProfile.setObjectName("pushButton_SaveProfile")
        self.pushButton_SaveProfile.setText("Save Profile")
        self.pushButton_DeleteProfile = QtWidgets.QPushButton(self)
        self.pushButton_DeleteProfile.setGeometry(QtCore.QRect(670, 120, 75, 23))
        self.pushButton_DeleteProfile.setObjectName("pushButton_DeleteProfile")
        self.pushButton_DeleteProfile.setText("Delete Profile")
        self.pushButton_UseProfile = QtWidgets.QPushButton(self)
        self.pushButton_UseProfile.setGeometry(QtCore.QRect(540, 210, 91, 31))
        self.pushButton_UseProfile.setObjectName("pushButton_UseProfile")
        self.pushButton_UseProfile.setText("Profil benutzen")
        self.pushButton_Back = QtWidgets.QPushButton(self)
        self.pushButton_Back.setGeometry(QtCore.QRect(670, 210, 81, 31))
        self.pushButton_Back.setObjectName("pushButton_Back")
        self.pushButton_Back.setText("Zurück")

        #Combobox befüllen mit vorhandenen Daten
        names = list(pd.read_csv("./EMS-Frontend/data/Gebäude_Profile.csv", usecols = [0], delimiter = ",", encoding='utf-8')["Name"])
        self.comboBox_SelectProfile.addItems(names)
        self.comboBox_SelectProfile.activated.connect(self.LoadProfile)
        self.pushButton_SaveProfile.clicked.connect(self.SaveProfile)
        self.pushButton_DeleteProfile.clicked.connect(self.DeleteProfile)
        self.pushButton_UseProfile.clicked.connect(self.UseProfile)

        self.li_inputWidgets = [self.doubleSpinBox_BGF,self.doubleSpinBox_GF,self.doubleSpinBox_cp_Gebaeude,
                                self.doubleSpinBox_Geschosshoehe]


    def SaveProfile(self):
        names = list(pd.read_csv("./EMS-Frontend/data/Gebäude_Profile.csv", usecols = [0], delimiter = ",", encoding='utf-8')["Name"])
        
        name = self.lineEdit_Profil.text()
        #Wenn nichts im Lineedit steht oder kein Radiobutton ausgewählt ist wird das Profil nicht gespeichert
        if name == "":
            return
   
        li_toSave = []
        li_toSave.append(name)

        for widget in self.li_inputWidgets:
            li_toSave.append(widget.value())

        for row in range(self.tableView_BauteilDaten.rowCount()):
            for column in range(4):
                item = self.tableView_BauteilDaten.item(row, column).text()
                if column == 0:
                    li_toSave.append(self.tableView_BauteilDaten.item(row, column).text())
                elif item != ""and is_number_tryexcept(item):
                    li_toSave.append(self.tableView_BauteilDaten.item(row, column).text())
                #Falls eine Zelle leer ist: Fehlermeldung
                else:
                    dlg = QMessageBox()
                    dlg.setWindowTitle("Fehler")
                    dlg.setText("Tabelle komplett ausfüllen bzw. auf Buchstaben kontrollieren")
                    dlg.exec()
                    return


        #Kontrolle ob ein Profil mit diesem Namen bereits existiert
        if name in names:
            self.DeleteProfile()
        #Neues Profil hinzufügen
        with open("./EMS-Frontend/data/Gebäude_Profile.csv",'a', newline='', encoding="utf-8") as f:
            writer = csv.writer(f, lineterminator='\n')
            writer.writerow(li_toSave)
        self.UpdateProfiles()
   
    def DeleteProfile(self):
        name = self.comboBox_SelectProfile.currentText()
        with open("./EMS-Frontend/data/Gebäude_Profile.csv", 'r', encoding="utf-8") as inp:
            lines = inp.readlines()
        with open("./EMS-Frontend/data/Gebäude_Profile.csv",'w', newline='', encoding="utf-8") as f:
            for line in lines:
                if line.split(",")[0] != name:
                    f.write(line)
        self.UpdateProfiles()
        

    def UpdateProfiles(self):
        names = list(pd.read_csv("./EMS-Frontend/data/Gebäude_Profile.csv", usecols = [0], delimiter = ",", encoding='utf-8')["Name"])
        self.comboBox_SelectProfile.clear() 
        self.comboBox_SelectProfile.addItems(names)        
        self.comboBox_SelectProfile.setCurrentText(self.lineEdit_Profil.text())
                   

    def LoadProfile(self):
     
        df = pd.read_csv("./EMS-Frontend/data/Gebäude_Profile.csv", delimiter = ",", encoding='utf-8')

        name = self.comboBox_SelectProfile.currentText()

        self.lineEdit_Profil.setText(name)
        values = df[df.values == name].values.flatten().tolist()
        df = df[df.values == name]

        for i,widget in enumerate(self.li_inputWidgets,1):
            try:
                widget.setValue(float(values[i]))
            except TypeError:
                widget.setValue(int(values[i]))

        #Counter sorgt dafür dass die ersten Wigdets geskippt werden
        counter = 5
        for row in range(self.tableView_BauteilDaten.rowCount()):
            for column in range(4):
                if column != 0:
                    self.tableView_BauteilDaten.setItem(row , column, QtWidgets.QTableWidgetItem(str(values[counter])))
                    counter += 1
                else:
                    #Falls Column 0 ist ist es ein Bauteilname und wird deswegen gesperrt
                    flags = QtCore.Qt.ItemFlags()
                    flags != QtCore.Qt.ItemIsEditable
                    item = QtWidgets.QTableWidgetItem(values[counter])
                    item.setFlags(flags)
                    self.tableView_BauteilDaten.setItem(row , column, item)
                    counter += 1


    def UseProfile(self):

        self.SaveProfile()

        df = pd.read_csv("./EMS-Frontend/data/Gebäude_Profile.csv", delimiter = ",", encoding='utf-8')
        name = self.comboBox_SelectProfile.currentText()
        data = df[df.values == name].values.flatten().tolist()

        from openpyxl import load_workbook
        wb = load_workbook(filename = "./EMS-Backend/data/building.xlsx")
        ws_params = wb["params"]
        ws_hull = wb["thermal_hull"]

        for row in range(1, 5):
            ws_params.cell(row=row+1,column=2).value = data[row]

        it = 5
        for row in range(2, 6): 
            for column in range(1,5):
                ws_hull.cell(row=row,column=column).value = data[it]
                it += 1
        wb.save("./EMS-Backend/data/building.xlsx")


def is_number_tryexcept(s):
    """ Returns True if string is a number. """
    try:
        float(s)
        return True
    except ValueError:
        return False

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

